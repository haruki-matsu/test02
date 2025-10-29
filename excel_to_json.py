import os
import json
import zipfile
from datetime import datetime, timedelta
from openpyxl import load_workbook


def main():
    # === 1️⃣ 設定 ===
    EXCEL_FILE = "Book1.xlsx"
    OUTPUT_JSON = "excel_analysis.json"
    OUTPUT_IMG_DIR = "output_images"

    # === 🆕 トラブル番号を指定 ===
    TROUBLE_ID = 21  # ← ここを変えるだけで "トラブル22" などにできる

    # 出力ディレクトリを生成
    TROUBLE_FOLDER = os.path.join(OUTPUT_IMG_DIR, f"トラブル{TROUBLE_ID}")
    os.makedirs(TROUBLE_FOLDER, exist_ok=True)

    # === 2️⃣ Excelの日付シリアル値を変換 ===
    def convert_excel_serial(value):
        """Excelの日付シリアル値を YYYY-MM-DD に変換"""
        try:
            base_date = datetime(1899, 12, 30)
            date_value = base_date + timedelta(days=int(value))
            if 1900 <= date_value.year <= 2100:
                return date_value.strftime("%Y-%m-%d")
            return value
        except Exception:
            return value

    # === 3️⃣ Excel本文データの抽出 ===
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.active

    text_cells = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                value = convert_excel_serial(cell.value)
                text_cells.append({
                    "type": "text",
                    "cell": cell.coordinate,
                    "value": str(value)
                })


    images_info = []
    image_counter = 1

    with zipfile.ZipFile(EXCEL_FILE, "r") as z:
        for file in z.namelist():
            if file.startswith("xl/media/"):
                new_name = f"トラブル{TROUBLE_ID}-{image_counter}.png"
                image_counter += 1

                extracted_path = os.path.join(TROUBLE_FOLDER, new_name)
                with z.open(file) as src, open(extracted_path, "wb") as dst:
                    dst.write(src.read())

                images_info.append({
                    "image_name": new_name,
                    "image_path": extracted_path.replace("\\", "/")
                })


    # === 5️⃣ 画像位置情報の取得 ===
    for image in sheet._images:
        try:
            if hasattr(image.anchor, "_from"):
                anchor_cell = f"{chr(65 + image.anchor._from.col)}{image.anchor._from.row + 1}"
            else:
                anchor_cell = str(image.anchor)
            idx = sheet._images.index(image)
            if idx < len(images_info):
                images_info[idx]["anchor_cell"] = anchor_cell
                images_info[idx]["width"] = image.width
                images_info[idx]["height"] = image.height
        except Exception as e:
            print(f"⚠️ 画像位置情報の取得エラー: {e}")


    # === 6️⃣ データ構造として保存 ===
    excel_data = {"text_cells": text_cells, "images": images_info}

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(excel_data, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Excel解析結果を {OUTPUT_JSON} に保存しました。")
    print(f"📁 画像フォルダ: {TROUBLE_FOLDER}")


if __name__ == "__main__":
    main()
